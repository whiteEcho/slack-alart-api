# encode: utf-8

from infrastructure.db_client import IClient
from openpyxl import load_workbook
from werkzeug.exceptions import BadRequest


class IncidentService:
    file_name = None
    file_path = None

    def __init__(self, file_name, file_path):
        self.file_name = file_name
        self.file_path = file_path

    def add(self):
        client = IClient()

        workbook = load_workbook(self.file_path)
        category, summary = self.__get_cell(workbook['表1　基本項目'])

        result = client.report_func(category, summary, self.file_name, self.file_path)
        workbook.close()

        return result

    def update(self, i_id):
        client = IClient()

        workbook = load_workbook(self.file_path)
        category, summary = self.__get_cell(workbook['表1　基本項目'])

        client.update_func(i_id, category, summary, self.file_name, self.file_path)
        workbook.close()

    def __get_cell(self, worksheet):
        # インシデント種別と概要の取得
        category, summary = worksheet['E4'].value, worksheet['E12'].value

        if category and summary:
            return category, self.__get_summary(summary)
        else:
            raise BadRequest

    @staticmethod
    def __get_summary(summary):
        return summary if len(summary) <= 30 else summary[:28] + '…'
